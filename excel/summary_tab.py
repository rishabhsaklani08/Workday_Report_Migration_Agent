"""
summary_tab.py — Create the Summary tab in the Excel workbook.

Columns updated to reflect the actual Workday RaaS field set
(Category, Output_Type, etc. instead of the fictional fields).
"""

import logging
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from config import COLORS
from utils.helpers import get_border

logger = logging.getLogger(__name__)


def create_summary_tab(wb, processed_data):
    """Build the Summary sheet as the first tab."""
    ws = wb.create_sheet("Summary", 0)

    # ── Title row ─────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "WORKDAY REPORT CATALOG — INDUSTRY SUMMARY"
    title_cell.font = Font(bold=True, size=16, color=COLORS["header_font"])
    title_cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Subtitle ──────────────────────────────────────────────
    ws.merge_cells("A2:I2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        f"  |  Source: Workday RaaS API"
        f"  |  Filter: AI_* tagged reports only"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])  # blank spacer

    # ── Headers ───────────────────────────────────────────────
    headers = [
        "Industry",
        "Total Reports",
        "Reports With Dependencies",
        "Reports Without Dependencies",
        "Dependency %",
        "Report Types",
        "Categories",
        "Data Sources",
        "Status",
    ]
    header_row = ws.max_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, size=11, color=COLORS["header_font"])
        cell.fill = PatternFill(
            fill_type="solid", fgColor=COLORS["summary_header"]
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = get_border()
    ws.row_dimensions[header_row].height = 30

    # ── Data rows ─────────────────────────────────────────────
    total_all = 0
    total_deps = 0
    for idx, (industry, reports) in enumerate(
        sorted(processed_data.items()), 1
    ):
        row_num = header_row + idx
        total = len(reports)
        with_deps = sum(
            1 for r in reports if r["Has_Dependencies"] == "✅ Yes"
        )
        without_deps = total - with_deps
        dep_pct = f"{(with_deps / total * 100):.1f}%" if total else "0%"

        types = ", ".join(
            sorted(
                {r["Report_Type"] for r in reports if r["Report_Type"] != "N/A"}
            )
        )
        categories = ", ".join(
            sorted(
                {r["Category"] for r in reports if r["Category"] != "N/A"}
            )
        )
        sources = len(
            {r["Data_Source"] for r in reports if r["Data_Source"] != "N/A"}
        )

        total_all += total
        total_deps += with_deps

        row_data = [
            industry,
            total,
            with_deps,
            without_deps,
            dep_pct,
            (types[:50] + "...") if len(types) > 50 else types,
            (categories[:50] + "...") if len(categories) > 50 else categories,
            f"{sources} unique source{'s' if sources != 1 else ''}",
            "✅ Active",
        ]
        row_fill = COLORS["alt_row_1"] if idx % 2 == 0 else COLORS["alt_row_2"]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.fill = PatternFill(fill_type="solid", fgColor=row_fill)
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )
            cell.border = get_border()

    # ── Totals row ────────────────────────────────────────────
    total_row = header_row + len(processed_data) + 1
    total_data = [
        "TOTAL",
        total_all,
        total_deps,
        total_all - total_deps,
        f"{(total_deps / total_all * 100):.1f}%" if total_all else "0%",
        "-",
        "-",
        "-",
        f"{len(processed_data)} Industries",
    ]
    for col, value in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="E3F2FD")
        cell.alignment = Alignment(horizontal="center")
        cell.border = get_border()

    # ── Column widths ─────────────────────────────────────────
    for letter, w in {
        "A": 25, "B": 15, "C": 25, "D": 28,
        "E": 15, "F": 25, "G": 25, "H": 20, "I": 15,
    }.items():
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = ws.cell(row=5, column=1)
    logger.info("✅ Summary tab created")
    return ws
