"""
industry_tab.py — Create individual industry tabs in the Excel workbook.

Columns match the actual Workday RaaS JSON structure:
    Report_Name, Report_Type, Output_Type, Category, Data_Source,
    Report_Owner, Report_Tag, Fields_Referenced_in_Report,
    Calculated_Fields_for_Report, Has_Dependencies, Dependency_Details,
    Created_On, Last_Updated
"""

import logging
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from config import COLORS, INDUSTRY_CONFIG
from utils.helpers import get_border

logger = logging.getLogger(__name__)

TAB_COLORS = INDUSTRY_CONFIG["tab_colors"]

# Column header definitions (13 data columns + row number = 14 total)
COLUMN_HEADERS = [
    "No.",                             # A
    "Report Name",                     # B
    "Report Type",                     # C
    "Output Type",                     # D
    "Category",                        # E
    "Data Source",                      # F
    "Report Owner",                    # G
    "Report Tag",                      # H
    "Fields Referenced",               # I
    "Calculated Fields",               # J
    "Has Dependencies",                # K
    "Dependency Details",              # L
    "Created On",                      # M
    "Last Updated",                    # N
]

COLUMN_WIDTHS = {
    "A": 6,   "B": 35,  "C": 14,  "D": 14,  "E": 18,
    "F": 28,  "G": 25,  "H": 22,  "I": 40,  "J": 40,
    "K": 18,  "L": 40,  "M": 22,  "N": 22,
}

# Columns that should be center-aligned (by 1-based index)
CENTER_COLS = {1, 3, 4, 11}

LAST_COL_LETTER = "N"
NUM_COLS = len(COLUMN_HEADERS)


def create_industry_tab(wb, industry, reports):
    """Create one industry sheet with all its AI-tagged reports."""

    # Sheet name: Excel max 31 chars, no special chars
    sheet_name = industry[:28].replace("/", "-").replace("\\", "-")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = TAB_COLORS.get(industry, "4A90D9")

    # ── Title row ─────────────────────────────────────────────
    ws.merge_cells(f"A1:{LAST_COL_LETTER}1")
    title = ws["A1"]
    title.value = f"WORKDAY REPORTS — {industry.upper()} INDUSTRY"
    title.font = Font(bold=True, size=14, color=COLORS["header_font"])
    title.fill = PatternFill(fill_type="solid", fgColor=COLORS["title_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Subtitle ──────────────────────────────────────────────
    deps_count = sum(1 for r in reports if r["Has_Dependencies"] == "✅ Yes")
    ws.merge_cells(f"A2:{LAST_COL_LETTER}2")
    ws["A2"].value = (
        f"Total Reports: {len(reports)}  |  "
        f"With Dependencies: {deps_count}  |  "
        f"Generated: {datetime.now().strftime('%d %B %Y')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.append([])  # blank separator row

    # ── Column headers ────────────────────────────────────────
    hdr_row = 4
    for col, h in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = Font(bold=True, size=11, color=COLORS["header_font"])
        cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = get_border()
    ws.row_dimensions[hdr_row].height = 35

    # ── Data rows ─────────────────────────────────────────────
    for idx, report in enumerate(reports, 1):
        row_num = hdr_row + idx
        has_deps = report["Has_Dependencies"] == "✅ Yes"

        row_data = [
            idx,
            report["Report_Name"],
            report["Report_Type"],
            report["Output_Type"],
            report["Category"],
            report["Data_Source"],
            report["Report_Owner"],
            report["Report_Tag"],
            report["Fields_Referenced_in_Report"],
            report["Calculated_Fields_for_Report"],
            report["Has_Dependencies"],
            report["Dependency_Details"],
            report["Created_On"],
            report["Last_Updated"],
        ]

        # Row background
        if has_deps:
            row_bg = "FFF9C4"  # soft yellow for dependency rows
        else:
            row_bg = COLORS["alt_row_1"] if idx % 2 == 0 else COLORS["alt_row_2"]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=str(value))
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
                horizontal="center" if col in CENTER_COLS else "left",
            )
            cell.border = get_border()
            cell.fill = PatternFill(fill_type="solid", fgColor=row_bg)

            # Special formatting for Has Dependencies column (K = col 11)
            if col == 11:
                if has_deps:
                    cell.fill = PatternFill(
                        fill_type="solid", fgColor=COLORS["dependency_yes"]
                    )
                    cell.font = Font(bold=True, color="1B5E20")
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FAFAFA")
                    cell.font = Font(color="999999")

        # Dynamic row height
        max_len = max(
            len(str(report.get("Fields_Referenced_in_Report", ""))),
            len(str(report.get("Calculated_Fields_for_Report", ""))),
            len(str(report.get("Dependency_Details", ""))),
        )
        ws.row_dimensions[row_num].height = max(20, min(80, max_len // 4))

    # ── Column widths ─────────────────────────────────────────
    for letter, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    # Freeze panes & auto-filter
    ws.freeze_panes = ws.cell(row=5, column=1)
    ws.auto_filter.ref = (
        f"A{hdr_row}:{LAST_COL_LETTER}{hdr_row + len(reports)}"
    )

    logger.info(f"✅ Created tab: {sheet_name} ({len(reports)} reports)")
    return ws
