"""
workday_agent.py — Consolidated Workday Report Catalog Agent.

This is a master script that contains all agent modules:
1. Environment & Configuration Loader
2. Workday RaaS API Client
3. JSON parser with AI_* tag filtering
4. Industry Tag Routing & Mapping
5. Dependency Checker (Calculated Fields Detector)
6. Excel Workbook Builder (Summary & Industry tabs)
7. Main Orchestration Flow
"""

import os
import sys
import re
import time
import logging
from datetime import datetime
from pathlib import Path

# Third-party library imports
import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────
# Console UTF-8 Settings for Windows stability
# ─────────────────────────────────────────────────────────────
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ─────────────────────────────────────────────────────────────
# Logging Setup
# ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("agent.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("WorkdayAgent")

# ─────────────────────────────────────────────────────────────
# SECTION 1: Environment & Configuration
# ─────────────────────────────────────────────────────────────
# Manual .env loader (removes external python-dotenv dependency)
_env_path = Path(__file__).resolve().parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text(encoding="utf-8").splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _key, _, _val = _line.partition("=")
            os.environ.setdefault(_key.strip(), _val.strip())

# Configuration Dictionaries
WORKDAY_CONFIG = {
    "raas_url": os.getenv(
        "WORKDAY_RAAS_URL",
        "https://[TENANT].workday.com/ccx/service/customreport2/[TENANT]/[USER]/[REPORT_NAME]"
    ),
    "isu_username": os.getenv("WORKDAY_ISU_USERNAME", "[ISU_USERNAME]"),
    "isu_password": os.getenv("WORKDAY_ISU_PASSWORD", "[ISU_PASSWORD]"),
    "timeout_seconds": int(os.getenv("WORKDAY_TIMEOUT", "60")),
    "retry_attempts": int(os.getenv("WORKDAY_RETRIES", "3")),
    "retry_delay_seconds": int(os.getenv("WORKDAY_RETRY_DELAY", "5")),
}

EXCEL_CONFIG = {
    "output_directory": os.getenv("OUTPUT_DIR", "./output"),
    "filename_prefix": "Workday_Report_Catalog",
    "include_timestamp": True,
}

COLORS = {
    "header_bg": "1E3A5F",       # Dark navy blue
    "header_font": "FFFFFF",     # White
    "alt_row_1": "F0F4F8",       # Light blue-gray
    "alt_row_2": "FFFFFF",       # White
    "dependency_yes": "C6EFCE",  # Light green
    "dependency_no": "FFCCCC",   # Light red
    "summary_header": "2E4057",  # Darker navy
    "tab_accent": "4A90D9",      # Blue accent
    "title_bg": "0D47A1",        # Deep blue title
    "border_color": "B0BEC5",    # Gray border
}

INDUSTRY_CONFIG = {
    "custom_mappings": {
        "AI_Healthcare": "Healthcare",
        "AI_HEALTHCARE": "Healthcare",
        "healthcare": "Healthcare",
        "AI_Insurance": "Insurance",
        "AI_INSURANCE": "Insurance",
        "insurance": "Insurance",
        "AI_Banking": "Banking",
        "AI_BANKING": "Banking",
        "AI_FinancialServices": "Financial Services",
        "AI_Finance": "Financial Services",
        "AI_Retail": "Retail",
        "AI_RETAIL": "Retail",
        "AI_Manufacturing": "Manufacturing",
        "AI_MANUFACTURING": "Manufacturing",
        "AI_Technology": "Technology",
        "AI_Tech": "Technology",
        "AI_Government": "Government",
        "AI_PublicSector": "Government",
        "AI_Education": "Education",
        "AI_EDUCATION": "Education",
        "AI_RealEstate": "Real Estate",
    },
    "tab_order": {
        "Summary": 0,
        "Healthcare": 1,
        "Insurance": 2,
        "Banking": 3,
        "Financial Services": 4,
        "Retail": 5,
        "Manufacturing": 6,
        "Technology": 7,
        "Government": 8,
        "Education": 9,
        "Real Estate": 10,
        "Other / Untagged": 999,
    },
    "tab_colors": {
        "Healthcare": "E91E63",
        "Insurance": "2196F3",
        "Banking": "4CAF50",
        "Financial Services": "FF9800",
        "Retail": "9C27B0",
        "Manufacturing": "795548",
        "Technology": "00BCD4",
        "Government": "607D8B",
        "Education": "FF5722",
        "Real Estate": "8BC34A",
        "Other / Untagged": "9E9E9E",
    },
}

UNMATCHED_TAB = "Other / Untagged"

DEPENDENCY_CONFIG = {
    "empty_values": ["", "null", "none", "n/a", "-", "no", "N/A"],
    "has_dependencies_text": "✅ Yes",
    "no_dependencies_text": "❌ No",
    "highlight_dependency_rows": True,
    "dependency_row_color": "FFF9C4",
}

# ─────────────────────────────────────────────────────────────
# SECTION 2: Helpers
# ─────────────────────────────────────────────────────────────
def get_border():
    """Return a standard thin cell border using the configured border color."""
    thin = Side(style="thin", color=COLORS["border_color"])
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ─────────────────────────────────────────────────────────────
# SECTION 3: Workday RaaS API Client
# ─────────────────────────────────────────────────────────────
def fetch_workday_report() -> dict:
    """Fetch report data from Workday via RaaS GET call."""
    url = WORKDAY_CONFIG["raas_url"]
    username = WORKDAY_CONFIG["isu_username"]
    password = WORKDAY_CONFIG["isu_password"]
    timeout = WORKDAY_CONFIG["timeout_seconds"]
    max_retries = WORKDAY_CONFIG["retry_attempts"]
    retry_delay = WORKDAY_CONFIG["retry_delay_seconds"]

    params = {"format": "json"}
    headers = {"Accept": "application/json", "Content-Type": "application/json"}
    last_exception = None

    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"📡 RaaS API call attempt {attempt}/{max_retries} → {url}")
            response = requests.get(
                url,
                auth=HTTPBasicAuth(username, password),
                params=params,
                headers=headers,
                timeout=timeout,
                verify=True,
            )

            if response.status_code == 200:
                logger.info("✅ RaaS call successful")
                return response.json()
            elif response.status_code == 401:
                raise Exception("Authentication failed (401). Check credentials in .env.")
            elif response.status_code == 403:
                raise Exception("Access denied (403). Check ISU permissions for report.")
            elif response.status_code == 404:
                raise Exception("Report not found (404). Verify WORKDAY_RAAS_URL.")
            elif response.status_code == 429:
                wait = retry_delay * attempt
                logger.warning(f"⏳ Rate limited (429). Retrying in {wait}s...")
                time.sleep(wait)
                continue
            else:
                raise Exception(f"API Error: {response.status_code} — {response.text[:500]}")
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            last_exception = exc
            logger.error(f"❌ Connection error/timeout on attempt {attempt}: {exc}")
            if attempt < max_retries:
                time.sleep(retry_delay)
        except Exception:
            raise

    raise Exception(f"All {max_retries} API attempts failed. Last error: {last_exception}")

# ─────────────────────────────────────────────────────────────
# SECTION 4: JSON Parser & AI_* Filter
# ─────────────────────────────────────────────────────────────
_AI_TAG_PATTERN = re.compile(r"AI_\w+", re.IGNORECASE)

def parse_json_response(json_data: dict) -> list[dict]:
    """Parse JSON and return only reports that have an AI_* Report_Tag."""
    all_reports = []
    if "Report_Entry" in json_data:
        all_reports = json_data["Report_Entry"]
    elif "report" in json_data:
        all_reports = json_data["report"]
    elif isinstance(json_data, list):
        all_reports = json_data
    else:
        for key, value in json_data.items():
            if isinstance(value, list) and len(value) > 0:
                logger.info(f"🔍 Auto-detected report array under key: '{key}'")
                all_reports = value
                break

    all_reports = [r for r in all_reports if isinstance(r, dict)]
    logger.info(f"📊 Total reports in raw JSON response: {len(all_reports)}")

    tagged_reports = []
    skipped = 0
    for report in all_reports:
        tag = report.get("Report_Tags") or report.get("Report_Tag") or ""
        if tag and _AI_TAG_PATTERN.search(str(tag)):
            tagged_reports.append(report)
        else:
            skipped += 1

    logger.info(f"🏷️  Reports with AI_* tag: {len(tagged_reports)} | Skipped (no AI tag): {skipped}")
    return tagged_reports

# ─────────────────────────────────────────────────────────────
# SECTION 5: Dependency Checker
# ─────────────────────────────────────────────────────────────
_EMPTY_VALUES = {v.strip().lower() for v in DEPENDENCY_CONFIG["empty_values"]}

def has_dependencies(calculated_fields) -> bool:
    """Return True if the report has calculated fields."""
    if calculated_fields is None:
        return False
    value = str(calculated_fields).strip()
    return value != "" and value.lower() not in _EMPTY_VALUES

def get_dependency_label(calculated_fields) -> str:
    return DEPENDENCY_CONFIG["has_dependencies_text"] if has_dependencies(calculated_fields) else DEPENDENCY_CONFIG["no_dependencies_text"]

def get_dependency_details(calculated_fields) -> str:
    return str(calculated_fields).strip() if has_dependencies(calculated_fields) else "None"

# ─────────────────────────────────────────────────────────────
# SECTION 6: Industry Router
# ─────────────────────────────────────────────────────────────
INDUSTRY_TAG_MAPPING = INDUSTRY_CONFIG["custom_mappings"]

def detect_industry(report_tag: str) -> str:
    """Resolve industry name from a single tag."""
    if not report_tag or str(report_tag).strip() == "":
        return UNMATCHED_TAB

    tag = str(report_tag).strip()

    # Case-insensitive exact & partial matching
    for mapped_tag, industry in INDUSTRY_TAG_MAPPING.items():
        if mapped_tag.lower() == tag.lower() or mapped_tag.lower() in tag.lower():
            return industry

    # Smart prefix extraction: AI_SmartCity -> Smart City
    if tag.upper().startswith("AI_"):
        industry_name = tag[3:].replace("_", " ").title()
        logger.info(f"🏷️  Auto-detected industry '{industry_name}' from tag '{tag}'")
        return industry_name

    return UNMATCHED_TAB

def handle_multiple_tags(report_tag: str) -> list[str]:
    """Split multi-value tags and resolve each tag."""
    if not report_tag or str(report_tag).strip() == "":
        return []

    report_tag = str(report_tag).strip()
    delimiters = [";", ",", "|", "/"]
    tags = [report_tag]

    for delimiter in delimiters:
        if delimiter in report_tag:
            tags = [t.strip() for t in report_tag.split(delimiter) if t.strip()]
            break

    industries = []
    for t in tags:
        industry = detect_industry(t)
        if industry != UNMATCHED_TAB and industry not in industries:
            industries.append(industry)
    return industries

# ─────────────────────────────────────────────────────────────
# SECTION 7: Summary Tab Generator
# ─────────────────────────────────────────────────────────────
def create_summary_tab(wb, processed_data):
    """Build the Summary tab with stats."""
    ws = wb.create_sheet("Summary", 0)

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "WORKDAY REPORT CATALOG — INDUSTRY SUMMARY"
    title_cell.font = Font(bold=True, size=16, color=COLORS["header_font"])
    title_cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle row
    ws.merge_cells("A2:I2")
    ws["A2"].value = (
        f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        f"  |  Source: Workday RaaS API  |  Filter: AI_* tagged reports only"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])  # Spacer

    # Headers
    headers = [
        "Industry", "Total Reports", "Reports With Dependencies",
        "Reports Without Dependencies", "Dependency %", "Report Types",
        "Categories", "Data Sources", "Status"
    ]
    header_row = ws.max_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, size=11, color=COLORS["header_font"])
        cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["summary_header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = get_border()
    ws.row_dimensions[header_row].height = 30

    total_all = 0
    total_deps = 0

    for idx, (industry, reports) in enumerate(sorted(processed_data.items()), 1):
        row_num = header_row + idx
        total = len(reports)
        with_deps = sum(1 for r in reports if r["Has_Dependencies"] == "✅ Yes")
        without_deps = total - with_deps
        dep_pct = f"{(with_deps / total * 100):.1f}%" if total else "0%"

        types = ", ".join(sorted({r["Report_Type"] for r in reports if r["Report_Type"] != "N/A"}))
        categories = ", ".join(sorted({r["Category"] for r in reports if r["Category"] != "N/A"}))
        sources = len({r["Data_Source"] for r in reports if r["Data_Source"] != "N/A"})

        total_all += total
        total_deps += with_deps

        row_data = [
            industry, total, with_deps, without_deps, dep_pct,
            (types[:50] + "...") if len(types) > 50 else types,
            (categories[:50] + "...") if len(categories) > 50 else categories,
            f"{sources} unique source{'s' if sources != 1 else ''}", "✅ Active"
        ]
        row_fill = COLORS["alt_row_1"] if idx % 2 == 0 else COLORS["alt_row_2"]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.fill = PatternFill(fill_type="solid", fgColor=row_fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = get_border()

    # Totals row
    total_row = header_row + len(processed_data) + 1
    total_data = [
        "TOTAL", total_all, total_deps, total_all - total_deps,
        f"{(total_deps / total_all * 100):.1f}%" if total_all else "0%",
        "-", "-", "-", f"{len(processed_data)} Industries"
    ]
    for col, value in enumerate(total_data, 1):
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="E3F2FD")
        cell.alignment = Alignment(horizontal="center")
        cell.border = get_border()

    # Widths
    for letter, w in {
        "A": 25, "B": 15, "C": 25, "D": 28, "E": 15, "F": 25, "G": 25, "H": 20, "I": 15
    }.items():
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = ws.cell(row=5, column=1)
    logger.info("✅ Summary tab created")
    return ws

# ─────────────────────────────────────────────────────────────
# SECTION 8: Industry Tab Generator
# ─────────────────────────────────────────────────────────────
COLUMN_HEADERS = [
    "No.", "Report Name", "Report Type", "Output Type", "Category",
    "Data Source", "Report Owner", "Report Tag", "Fields Referenced",
    "Calculated Fields", "Has Dependencies", "Dependency Details",
    "Created On", "Last Updated"
]

COLUMN_WIDTHS = {
    "A": 6, "B": 35, "C": 14, "D": 14, "E": 18, "F": 28, "G": 25,
    "H": 22, "I": 40, "J": 40, "K": 18, "L": 40, "M": 22, "N": 22
}

CENTER_COLS = {1, 3, 4, 11}
LAST_COL_LETTER = "N"

def create_industry_tab(wb, industry, reports):
    """Build a sheet for a specific industry tab."""
    sheet_name = industry[:28].replace("/", "-").replace("\\", "-")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = INDUSTRY_CONFIG["tab_colors"].get(industry, "4A90D9")

    # Title row
    ws.merge_cells(f"A1:{LAST_COL_LETTER}1")
    title = ws["A1"]
    title.value = f"WORKDAY REPORTS — {industry.upper()} INDUSTRY"
    title.font = Font(bold=True, size=14, color=COLORS["header_font"])
    title.fill = PatternFill(fill_type="solid", fgColor=COLORS["title_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    deps_count = sum(1 for r in reports if r["Has_Dependencies"] == "✅ Yes")
    ws.merge_cells(f"A2:{LAST_COL_LETTER}2")
    ws["A2"].value = (
        f"Total Reports: {len(reports)}  |  With Dependencies: {deps_count}  |  "
        f"Generated: {datetime.now().strftime('%d %B %Y')}"
    )
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.append([])  # Spacer

    # Headers
    hdr_row = 4
    for col, h in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = Font(bold=True, size=11, color=COLORS["header_font"])
        cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = get_border()
    ws.row_dimensions[hdr_row].height = 35

    # Data
    for idx, report in enumerate(reports, 1):
        row_num = hdr_row + idx
        has_deps = report["Has_Dependencies"] == "✅ Yes"

        row_data = [
            idx, report["Report_Name"], report["Report_Type"], report["Output_Type"],
            report["Category"], report["Data_Source"], report["Report_Owner"],
            report["Report_Tag"], report["Fields_Referenced_in_Report"],
            report["Calculated_Fields_for_Report"], report["Has_Dependencies"],
            report["Dependency_Details"], report["Created_On"], report["Last_Updated"]
        ]

        row_bg = "FFF9C4" if has_deps else (COLORS["alt_row_1"] if idx % 2 == 0 else COLORS["alt_row_2"])

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=str(value))
            cell.alignment = Alignment(
                vertical="top", wrap_text=True,
                horizontal="center" if col in CENTER_COLS else "left"
            )
            cell.border = get_border()
            cell.fill = PatternFill(fill_type="solid", fgColor=row_bg)

            # Special column formatting for dependencies (col 11 / K)
            if col == 11:
                if has_deps:
                    cell.fill = PatternFill(fill_type="solid", fgColor=COLORS["dependency_yes"])
                    cell.font = Font(bold=True, color="1B5E20")
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor="FAFAFA")
                    cell.font = Font(color="999999")

        max_len = max(
            len(str(report.get("Fields_Referenced_in_Report", ""))),
            len(str(report.get("Calculated_Fields_for_Report", ""))),
            len(str(report.get("Dependency_Details", "")))
        )
        ws.row_dimensions[row_num].height = max(20, min(80, max_len // 4))

    for letter, w in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = ws.cell(row=5, column=1)
    ws.auto_filter.ref = f"A{hdr_row}:{LAST_COL_LETTER}{hdr_row + len(reports)}"

    logger.info(f"✅ Created tab: {sheet_name} ({len(reports)} reports)")
    return ws

# ─────────────────────────────────────────────────────────────
# SECTION 9: Excel Workbook Builder
# ─────────────────────────────────────────────────────────────
TAB_ORDER = INDUSTRY_CONFIG["tab_order"]

def create_excel_workbook(processed_data: dict, output_path: str) -> str:
    """Create the full workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default active sheet

    industries = sorted(processed_data.keys(), key=lambda x: TAB_ORDER.get(x, 500))

    # Add Summary first
    create_summary_tab(wb, processed_data)

    # Add industry tabs
    for ind in industries:
        create_industry_tab(wb, ind, processed_data[ind])

    wb.save(output_path)
    logger.info(f"✅ Excel file saved: {output_path}")
    return output_path

# ─────────────────────────────────────────────────────────────
# SECTION 10: Orchestrator & Sample Data
# ─────────────────────────────────────────────────────────────
def process_reports(raw_reports: list[dict]) -> dict:
    """Process & route raw API reports into industry tab categories."""
    processed = {}
    for report in raw_reports:
        report_data = {
            "Report_Name": report.get("Report_Name", "N/A"),
            "Report_Type": report.get("Report_Type", "N/A"),
            "Output_Type": report.get("Output_Type", "N/A"),
            "Category": report.get("Category", "N/A"),
            "Data_Source": report.get("Data_Source", "N/A"),
            "Report_Owner": report.get("Report_Owner", "N/A"),
            "Report_Tag": report.get("Report_Tags") or report.get("Report_Tag") or "",
            "Fields_Referenced_in_Report": report.get("Fields_Referenced_in_Report", ""),
            "Calculated_Fields_for_Report": report.get("Calculated_Fields_for_Report", ""),
            "Created_On": report.get("Created_On", "N/A"),
            "Last_Updated": report.get("Last_Updated", "N/A"),
            "Shared": report.get("Shared", "N/A"),
            "Worklet": report.get("Worklet", "N/A"),
        }

        calc_fields = report_data["Calculated_Fields_for_Report"]
        report_data["Has_Dependencies"] = get_dependency_label(calc_fields)
        report_data["Dependency_Details"] = get_dependency_details(calc_fields)

        industries = handle_multiple_tags(report_data["Report_Tag"])
        if not industries:
            continue

        for industry in industries:
            processed.setdefault(industry, []).append(report_data)

    return processed

def get_sample_data() -> dict:
    """Mock dataset mirroring the actual Workday JSON format."""
    return {
        "Report_Entry": [
            {
                "Report_Name": "<tenant> Patient Demographics Report",
                "Report_Type": "Advanced",
                "Output_Type": "Table",
                "Category": "Patient Data",
                "Data_Source": "Workers for HCM Reporting",
                "Report_Owner": "admin-user / Admin User",
                "Report_Tag": "AI_Healthcare",
                "Fields_Referenced_in_Report": "Worker; Age; Worker Types",
                "Calculated_Fields_for_Report": "PTN_<tenant>_AGE_CALC; PTN_<tenant>_LOS_CALC",
                "Created_On": "2026-05-01T10:30:00.000-07:00",
                "Last_Updated": "2026-05-10T14:22:00.000-07:00",
                "Shared": "1",
                "Worklet": "0",
            },
            {
                "Report_Name": "<tenant> Claims Summary Q1",
                "Report_Type": "Advanced",
                "Output_Type": "Table",
                "Category": "Claims Data",
                "Data_Source": "All Active Workers",
                "Report_Owner": "claims-user / Claims Analyst",
                "Report_Tag": "AI_Insurance",
                "Fields_Referenced_in_Report": "Claim_ID; Policy_Number; Amount",
                "Calculated_Fields_for_Report": "",
                "Created_On": "2026-03-01T12:00:00.000-07:00",
                "Last_Updated": "2026-05-08T16:30:00.000-07:00",
                "Shared": "1",
                "Worklet": "0",
            },
            {
                "Report_Name": "<tenant> Daily Transactions",
                "Report_Type": "Composite",
                "Output_Type": "Table",
                "Category": "Transaction Data",
                "Data_Source": "Transaction Ledger",
                "Report_Owner": "banking-user / Banking Ops",
                "Report_Tag": "AI_Banking",
                "Fields_Referenced_in_Report": "TXN_ID; Amount; Date",
                "Calculated_Fields_for_Report": "PTN_<tenant>_NET_BALANCE",
                "Created_On": "2026-01-10T07:00:00.000-07:00",
                "Last_Updated": "2026-05-13T08:20:00.000-07:00",
                "Shared": "1",
                "Worklet": "0",
            },
            {
                "Report_Name": "<tenant> Cross-Industry Health-Insurance",
                "Report_Type": "Composite",
                "Output_Type": "Table",
                "Category": "Cross-Industry",
                "Data_Source": "Data Warehouse",
                "Report_Owner": "admin-user / Admin User",
                "Report_Tag": "AI_Healthcare; AI_Insurance",
                "Fields_Referenced_in_Report": "Patient_ID; Claim_ID; Amount",
                "Calculated_Fields_for_Report": "PTN_<tenant>_COMBINED_COST",
                "Created_On": "2026-05-05T10:00:00.000-07:00",
                "Last_Updated": "2026-05-13T12:00:00.000-07:00",
                "Shared": "0",
                "Worklet": "0",
            },
            # Untagged report (should be skipped)
            {
                "Report_Name": "<snguvvala-trn> Worker Details and Dependents",
                "Report_Type": "Advanced",
                "Output_Type": "Table",
                "Category": "Worker Data",
                "Data_Source": "Workers for HCM Reporting",
                "Report_Owner": "snguvvala-trn / siva nagi guvvala",
                "Fields_Referenced_in_Report": "Worker; Age; Worker Types",
                "Calculated_Fields_for_Report": "PTN_<snguvvala-trn>_LRV_EXTRACT_OLDEST",
                "Created_On": "2026-05-11T04:43:40.048-07:00",
                "Last_Updated": "2026-05-13T02:05:39.208-07:00",
                "Shared": "0",
                "Worklet": "0",
            }
        ]
    }

def run_agent(output_path=None, use_sample_data=False):
    """Main pipeline execution."""
    print("🤖 Workday Report Catalog Agent Starting (Consolidated Master File)...")
    print("=" * 70)

    # 1. Fetch
    if use_sample_data:
        logger.info("Using sample data (offline mode)")
        raw_data = get_sample_data()
    else:
        raw_data = fetch_workday_report()

    # 2. Parse & Filter
    reports = parse_json_response(raw_data)
    if not reports:
        logger.warning("⚠️ No AI-tagged reports found in response. Check Report_Tag values.")
        return None

    # 3. Categorize & Route
    processed = process_reports(reports)
    if not processed:
        logger.warning("⚠️ No reports matched configured industries.")
        return None

    # 4. Print Summary
    print("\n📊 Categorization Summary:")
    print("-" * 50)
    for industry, rpts in sorted(processed.items()):
        deps = sum(1 for r in rpts if r["Has_Dependencies"] == "✅ Yes")
        print(f"  {industry:<25} {len(rpts):>3} reports ({deps} with dependencies)")
    print("-" * 50)
    print(f"  {'TOTAL':<25} {sum(len(r) for r in processed.values()):>3} reports")

    # 5. Generate Excel workbook
    output_dir = EXCEL_CONFIG["output_directory"]
    os.makedirs(output_dir, exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = EXCEL_CONFIG["filename_prefix"]
        output_path = os.path.join(output_dir, f"{prefix}_{timestamp}.xlsx")

    create_excel_workbook(processed, output_path)

    print("\n" + "=" * 70)
    print("✅ Agent completed successfully!")
    print(f"📁 Output file: {os.path.abspath(output_path)}")
    print(f"📋 Total AI-tagged reports processed: {len(reports)}")
    print("=" * 70)
    return output_path

# ─────────────────────────────────────────────────────────────
# CLI Entrypoint
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    use_sample = "--sample" in sys.argv
    run_agent(use_sample_data=use_sample)
